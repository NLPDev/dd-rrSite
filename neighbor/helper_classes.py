from django.http import HttpResponse

__author__ = 'nathanlebert'

from datetime import date
from calendar import monthrange, IllegalMonthError
from django import forms
from django.conf import settings
from django.utils.translation import ugettext_lazy as _


CREDIT_CARD_RE = r'^(?:4[0-9]{12}(?:[0-9]{3})?|5[1-5][0-9]{14}|6(?:011|5[0-9][0-9])[0-9]{12}|3[47][0-9]{13}|3(?:0[0-5]|[68][0-9])[0-9]{11}|(?:2131|1800|35\\d{3})\d{11})$'
MONTH_FORMAT = getattr(settings, 'MONTH_FORMAT', '%b')
VERIFICATION_VALUE_RE = r'^([0-9]{3,4})$'
# Taken from the django-creditcard app.

class ExpiryDateWidget(forms.MultiWidget):
    """
    Widget containing two select boxes for selecting the month and year.
    """

    def decompress(self, value):
        return [value.month, value.year] if value else [None, None]

    def format_output(self, rendered_widgets):
        return u'<div class="expirydatefield col-xs-12 col-sm-12 col-md-6">%s</div><div class="expirydatefield col-xs-12 col-sm-12 col-md-6">%s</div>' % (rendered_widgets[0], rendered_widgets[1])


class ExpiryDateField(forms.MultiValueField):
    """
    Form field that validates credit card expiry dates.
    """

    default_error_messages = {
        'invalid_month': _(u'Please enter a valid month.'),
        'invalid_year': _(u'Please enter a valid year.'),
        'date_passed': _(u'This expiry date has passed.'),
    }

    def __init__(self, *args, **kwargs):
        today = date.today()
        error_messages = self.default_error_messages.copy()
        if 'error_messages' in kwargs:
            error_messages.update(kwargs['error_messages'])
        if 'initial' not in kwargs:
            # Set default expiry date based on current month and year
            kwargs['initial'] = today
        months = [(x, '%02d (%s)' % (x, date(2000, x, 1).strftime(MONTH_FORMAT))) for x in xrange(1, 13)]
        years = [(x, x) for x in xrange(today.year, today.year + 15)]
        fields = (
            forms.ChoiceField(choices=months, error_messages={'invalid': error_messages['invalid_month']}),
            forms.ChoiceField(choices=years, error_messages={'invalid': error_messages['invalid_year']}),
        )
        super(ExpiryDateField, self).__init__(fields, *args, **kwargs)
        self.widget = ExpiryDateWidget(widgets=[fields[0].widget, fields[1].widget])

    def clean(self, value):
        expiry_date = super(ExpiryDateField, self).clean(value)
        if date.today() > expiry_date:
            raise forms.ValidationError(self.error_messages['date_passed'])
        return expiry_date

    def compress(self, data_list):
        if data_list:
            try:
                month = int(data_list[0])
            except (ValueError, TypeError):
                raise forms.ValidationError(self.error_messages['invalid_month'])
            try:
                year = int(data_list[1])
            except (ValueError, TypeError):
                raise forms.ValidationError(self.error_messages['invalid_year'])
            try:
                day = monthrange(year, month)[1] # last day of the month
            except IllegalMonthError:
                raise forms.ValidationError(self.error_messages['invalid_month'])
            except ValueError:
                raise forms.ValidationError(self.error_messages['invalid_year'])
            return date(year, month, day)
        return None



def export_xlsx(modeladmin, request, queryset):
    import openpyxl
    from openpyxl.cell import get_column_letter
    from StringIO import StringIO
    import datetime
    from django.utils.encoding import smart_str

    output = StringIO()
    wb = openpyxl.Workbook()
    ws = wb.get_active_sheet()
    ws.title = "Purchase Report"

    row_num = 0

    columns = [
        (u"ID", 10),
        (u"Date", 20),
        (u"Amount", 10),
        (u"First Name", 30),
        (u"Last Name", 30),
        (u"Transaction Type", 20),
        (u"Community", 30),
    ]

    for col_num in xrange(len(columns)):
        c = ws.cell(row=row_num + 1, column=col_num + 1)
        c.value = columns[col_num][0]
        #c.style.font.bold = True
        # set column width
        ws.column_dimensions[get_column_letter(col_num+1)].width = columns[col_num][1]

    for obj in queryset:
        row_num += 1
        row = [
            smart_str(obj.pk),
            smart_str(obj.date),
            smart_str(obj.amount),
            smart_str(obj.wallet.neighbor.first_name),
            smart_str(obj.wallet.neighbor.last_name),
            smart_str(obj.transaction_type),
            smart_str(obj.wallet.neighbor.community),
        ]
        for col_num in xrange(len(row)):
            c = ws.cell(row=row_num + 1, column=col_num + 1)
            c.value = row[col_num]
            #c.style.alignment.wrap_text = True

    wb.save(output)

    output.seek(0)
    response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=PurchaseReport_%s.xlsx' % datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    return response
export_xlsx.short_description = u"Export XLSX"


def export_csv(modeladmin, request, queryset):
    import csv
    from django.utils.encoding import smart_str
    from StringIO import StringIO
    import datetime

    output = StringIO()
    writer = csv.writer(output, csv.excel)
    output.write(u'\ufeff'.encode('utf8')) # BOM (optional...Excel needs it to open UTF-8 file properly)
    writer.writerow([
        smart_str(u"ID"),
        smart_str(u"Date"),
        smart_str(u"Amount"),
        smart_str(u"First Name"),
        smart_str(u"Last Name"),
        smart_str(u"Transaction Type"),
        smart_str(u"Community"),
    ])
    for obj in queryset:

        writer.writerow([
            smart_str(obj.pk),
            smart_str(obj.date),
            smart_str(obj.amount),
            smart_str(obj.wallet.neighbor.first_name),
            smart_str(obj.wallet.neighbor.last_name),
            smart_str(obj.transaction_type),
            smart_str(obj.wallet.neighbor.community),
        ])

    output.seek(0)
    response = HttpResponse(output.read(), content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename=PurchaseReport_%s.csv' % datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    return response
export_csv.short_description = u"Export CSV"